"""Import the TCH MIS workbook into the database.

Usage:
    python manage.py import_excel /path/to/workbook.xlsx [--wipe]

The CommercialDeal table is the source of truth — Current Overview is derived
from it, so we deliberately do not import the Current Overview sheet's
hand-entered totals (those would just shadow the live aggregation).

Workbook variants:
- "Historical" (sourceOriginalHistorical.xlsx): Current Overview is a creator
  roster (profile, location, ops mgr); Commercial Tracking has no Deliverables
  column; ships with Drop-offs sheet.
- "FY 26-27" (TCH MIS FY 26-27.xlsx): Current Overview is the totals matrix;
  Commercial Tracking has a Deliverables column at index 12; no Drop-offs.

We auto-detect by looking at the Commercial Tracking header at column M (index 12).
"""

from __future__ import annotations

import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from dateutil import parser as dateparser
from django.core.management.base import BaseCommand
from django.db import transaction

import openpyxl

from tch.models import (
    Campaign, Creator, ContractingCompliance, CommercialDeal, EmployeeWeeklyReport, DropOff,
)


# Map "raw name in Commercial Tracking" -> canonical pipeline name.
# The historical file has typos and spelling drift the pipeline doesn't.
NAME_ALIASES = {
    "monica d'zousa": "Monica D'souza",
    "monica dsouza": "Monica D'souza",
    "monica d'souza": "Monica D'souza",
    "prateik jain": "Prateik Jaiin",
    "shreeradhe": "Shree Radhe",
    "shree radhe": "Shree Radhe",
    "kaam bhari": "Kaam Bhaari",
    "kaam bhaari": "Kaam Bhaari",
    "antara asthana": "Antara Asthana",
    "saili satwe": "Saili Satwe",
}


def _to_date(v):
    """Parse a cell value into a date.

    The source workbook is Indian (DD/MM/YYYY), but Excel often auto-converts
    text dates into native date cells using whatever locale the editing client
    had. When openpyxl gives us a real datetime back, we can no longer see the
    original string, so we apply this heuristic:

      If both day and month are <= 12, the original string was ambiguous.
      Swap them — for an Indian dataset that matches the typed intent.

    If day > 12, the date was typed unambiguously (e.g. 25/03/2026) — keep.
    Plain-text dates from openpyxl go through dateparser with dayfirst=True.
    """
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        d = v.date()
        return _swap_if_ambiguous(d)
    if isinstance(v, date):
        return _swap_if_ambiguous(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return dateparser.parse(s, dayfirst=True, fuzzy=True).date()
    except (ValueError, TypeError, OverflowError):
        return None


def _swap_if_ambiguous(d: date) -> date:
    """Swap day/month if both are <= 12 (locale-mismatch fix for Indian DD/MM)."""
    if d.day <= 12 and d.month <= 12:
        try:
            return date(d.year, d.day, d.month)
        except ValueError:
            return d
    return d


def _to_decimal(v) -> Decimal:
    if v is None or v == '':
        return Decimal('0')
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return Decimal('0')
    s = str(v).strip().replace(',', '').replace('₹', '')
    if not s or s.upper() == 'N/A':
        return Decimal('0')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _to_yn(v) -> str:
    if v is None:
        return ''
    s = str(v).strip().upper()
    return s if s in ('Y', 'N') else ''


def _to_pct(v) -> Decimal:
    """Accept 0.2, '20%', '20', 'NA', etc. Return fractional decimal (0..1)."""
    if v is None or v == '':
        return Decimal('0')
    if isinstance(v, (int, float, Decimal)):
        d = Decimal(str(v))
        return d if d <= 1 else d / Decimal('100')
    s = str(v).strip()
    if not s or s.upper() in ('NA', 'N/A'):
        return Decimal('0')
    if s.endswith('%'):
        try:
            return Decimal(s[:-1]) / Decimal('100')
        except InvalidOperation:
            return Decimal('0')
    try:
        d = Decimal(s)
        return d if d <= 1 else d / Decimal('100')
    except InvalidOperation:
        return Decimal('0')


def _norm_name(name: str) -> str:
    if not name:
        return ''
    # Apply alias map; otherwise just strip whitespace
    key = re.sub(r"\s+", ' ', name.strip().lower())
    return NAME_ALIASES.get(key, name.strip())


def _resolve_creator(name: str, default_relationship: str = 'NonTCH') -> Creator | None:
    canonical = _norm_name(name)
    if not canonical:
        return None
    creator = Creator.objects.filter(name__iexact=canonical).first()
    if creator:
        return creator
    return Creator.objects.create(name=canonical, relationship=default_relationship)


def _resolve_campaign(name: str, brand: str = '') -> Campaign | None:
    """Match a campaign name (case-insensitive, whitespace-collapsed) to an
    existing Campaign, creating it on first sight."""
    canonical = ' '.join((name or '').split())[:255]
    if not canonical:
        return None
    campaign = Campaign.objects.filter(name__iexact=canonical).first()
    if campaign:
        return campaign
    return Campaign.objects.create(name=canonical, brand=(brand or '')[:200])


class Command(BaseCommand):
    help = "Import the TCH MIS workbook into the DB."

    def add_arguments(self, parser):
        parser.add_argument('path', type=str)
        parser.add_argument('--wipe', action='store_true', help='Delete existing rows first')

    @transaction.atomic
    def handle(self, *args, path: str, wipe: bool = False, **kwargs):
        wb = openpyxl.load_workbook(path, data_only=True)
        if wipe:
            self.stdout.write('Wiping existing rows...')
            CommercialDeal.objects.all().delete()
            Campaign.objects.all().delete()
            ContractingCompliance.objects.all().delete()
            EmployeeWeeklyReport.objects.all().delete()
            DropOff.objects.all().delete()
            Creator.objects.all().delete()

        # Always: pipeline first (creates the master), then enrich from Current Overview
        # roster (if present), then contracting, then deals, then employees, then drops.
        if 'Creator Pipeline' in wb.sheetnames:
            self._import_creators(wb['Creator Pipeline'])
        if 'Current Overview' in wb.sheetnames:
            self._enrich_from_roster(wb['Current Overview'])
        if 'Contracting & Compliance' in wb.sheetnames:
            self._import_contracting(wb['Contracting & Compliance'])
        if 'Commercial Tracking' in wb.sheetnames:
            self._import_deals(wb['Commercial Tracking'])
        if 'Employee-Talent' in wb.sheetnames:
            self._import_employee_reports(wb['Employee-Talent'])
        if 'Drop-offs' in wb.sheetnames:
            self._import_dropoffs(wb['Drop-offs'])

        self.stdout.write(self.style.SUCCESS('Import complete.'))

    # ---- per-sheet importers ----

    def _import_creators(self, ws):
        """Creator Pipeline: cols A=S No, B=Name, C=Category, D=Source, E=Stage,
        F=Exclusive/Friend, G=DOJ, H=Notes (historical only)."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[1]
            if not name_raw:
                continue
            name = _norm_name(str(name_raw))
            category = (row[2] or '').strip() if row[2] else ''
            source_raw = (row[3] or '').strip().upper() if row[3] else ''
            source = source_raw if source_raw in ('EMW', 'TCH') else ('OTHER' if source_raw else '')
            stage_raw = (row[4] or '').strip() if row[4] else ''
            stage = stage_raw if stage_raw in ('Lead', 'Discussing', 'Closed', 'Dropped') else ''
            rel_raw = (row[5] or '').strip().lower() if row[5] else ''
            relationship = (
                'Exclusive' if 'exclusive' in rel_raw else
                'Friend' if 'friend' in rel_raw else
                'Friend'
            )
            # Stage=Dropped overrides into the "Dropping out soon" bucket so the
            # Overview correctly attributes their historical billing.
            if stage == 'Dropped':
                relationship = 'Dropping'
            doj = _to_date(row[6])
            doj_note = '' if doj else (str(row[6]).strip() if row[6] else '')
            notes = str(row[7]).strip() if len(row) > 7 and row[7] else ''

            Creator.objects.update_or_create(
                name=name,
                defaults={
                    'category': category,
                    'source': source,
                    'stage': stage,
                    'relationship': relationship,
                    'doj': doj,
                    'doj_note': doj_note[:120],
                    'notes': notes,
                },
            )
            count += 1
        self.stdout.write(f'  creators (pipeline): {count}')

    def _enrich_from_roster(self, ws):
        """In the historical workbook, 'Current Overview' is actually a creator
        roster: SR | Name | Profile URL | Location | Category Affinity | Ops Manager [| 'E' marker].
        In the FY 26-27 workbook, it's a totals matrix with no useful per-creator
        data — that case is detected by the absence of Profile URLs."""
        enriched = 0
        # Header row is row 3 in this layout
        for row in ws.iter_rows(min_row=4, values_only=True):
            name_raw = row[1] if len(row) > 1 else None
            profile = row[2] if len(row) > 2 else None
            if not name_raw or not profile:
                continue
            if not isinstance(profile, str) or not profile.startswith('http'):
                continue
            name = _norm_name(str(name_raw))
            location = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            category = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            ops = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            e_marker = str(row[6]).strip().upper() if len(row) > 6 and row[6] else ''

            creator = Creator.objects.filter(name__iexact=name).first()
            if not creator:
                creator = Creator.objects.create(
                    name=name,
                    relationship='Exclusive' if e_marker == 'E' else 'Friend',
                )
            creator.profile_url = profile.strip()[:400]
            if location:
                creator.location = location[:80]
            if category and not creator.category:
                creator.category = category[:200]
            if ops:
                creator.ops_manager = ops[:80]
            if e_marker == 'E' and creator.relationship not in ('Exclusive', 'Dropping'):
                creator.relationship = 'Exclusive'
            creator.save()
            enriched += 1
        self.stdout.write(f'  creators (enriched from roster): {enriched}')

    def _import_contracting(self, ws):
        """Header row 3, data row 4+. Cols: A=S No, B=Name, C=FinalMtg, D=AgrSent,
        E=AgrSigned, F=BankVerified, G=TimeToSign, H=RenewalDate."""
        count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            name_raw = row[1]
            if not name_raw:
                continue
            name = _norm_name(str(name_raw))
            creator, _ = Creator.objects.get_or_create(
                name=name, defaults={'relationship': 'Friend'}
            )
            renewal = _to_date(row[7])
            ContractingCompliance.objects.update_or_create(
                creator=creator,
                defaults={
                    'final_meeting': _to_yn(row[2]),
                    'agreement_sent': _to_yn(row[3]),
                    'agreement_signed': _to_yn(row[4]),
                    'bank_verified': _to_yn(row[5]),
                    'time_to_sign': str(row[6] or '')[:120],
                    'renewal_date': renewal,
                    'renewal_note': '' if renewal else str(row[7] or '')[:120],
                },
            )
            count += 1
        self.stdout.write(f'  contracting: {count}')

    def _import_deals(self, ws):
        """Detect column layout by reading the header row (row 3 in the newer
        files, but tolerate any of the three known layouts).

        Historical layout:
          A Conf | B EInvDate | C Creator | D Commission | E Dir | F Total | G %
          H AgcyINR | I CreatorFee | J BillEntity | K Brand | L Campaign
          M RO# | N Over | O InvRecv | P PayClr | Q EInv# | R PayRecv | S Comments

        FY 26-27 layout (extra Deliverables col before RO#):
          ... L Campaign | M Deliverables | N RO# | O Over | P InvRecv | Q PayClr |
          R EInv# | S PayRecv | T Comments

        FY 26-27 "invoicing tracker" layout (no Commission col; detected by
        Deliverables sitting at column L):
          A Conf | B EInvDate | C Creator | D Dir | E Total | F % | G AgcyINR
          H CreatorFee | I BillEntity | J Brand | K Campaign | L Deliverables
          M RO# | N Over | O InvRecv | P PayClr | Q EInv# | R PayRecv | S Comments
          Multi-creator campaigns appear as RO sub-rows (TCH/x/152/A, /B, …)
          whose brand/campaign cells are blank — they inherit from the group's
          first row so all sub-rows link to the same Campaign.
        """
        header_L = str(ws.cell(row=3, column=12).value or '')   # column L
        header_M = str(ws.cell(row=3, column=13).value or '')   # column M
        if 'deliverable' in header_L.lower():
            layout = 'tracker'
        elif 'deliverable' in header_M.lower():
            layout = 'fy2627'
        else:
            layout = 'historical'

        # Already-imported deals (by RO # / E-Invoice #) are skipped so the
        # command can be re-run on overlapping workbooks without duplicating.
        def _key(v):
            return ' '.join(str(v or '').split()).lower()

        existing_ro = {
            _key(v) for v in CommercialDeal.objects.values_list('ro_number', flat=True)
            if v and _key(v) not in ('', 'na')
        }
        existing_einv = {
            _key(v) for v in CommercialDeal.objects.values_list('e_invoice_number', flat=True) if v
        }

        # RO numbers like "TCH/2526/152/A" / "TCH/2526/155 B" mark sub-rows of
        # one multi-creator campaign; the base ("TCH/2526/152") groups them.
        ro_base_re = re.compile(r'^(.*?\d+)\s*[/ ]\s*([A-Za-z])$')

        parsed = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not any(row):
                continue

            def cell(i):
                return row[i] if len(row) > i else None

            if layout == 'tracker':
                d = {
                    'confirmation': _to_date(cell(0)),
                    'invoice_date': _to_date(cell(1)),
                    'cname_raw': str(cell(2) or '').strip(),
                    'agency_commission': '',
                    'direction_raw': str(cell(3) or ''),
                    'total_fee': _to_decimal(cell(4)),
                    'agency_pct': _to_pct(cell(5)),
                    'agency_inr': _to_decimal(cell(6)),
                    'creator_fee': _to_decimal(cell(7)),
                    'billing_entity': str(cell(8) or '')[:120],
                    'brand': str(cell(9) or '').strip()[:200],
                    'campaign_name': str(cell(10) or ''),
                    'deliverables': str(cell(11) or '')[:255],
                    'ro_number': str(cell(12) or '')[:80],
                    'campaign_over': _to_yn(cell(13)),
                    'invoice_received': _to_yn(cell(14)),
                    'payment_cleared': _to_yn(cell(15)),
                    'e_invoice_no': str(cell(16) or '')[:80],
                    'payment_received': _to_yn(cell(17)),
                    'comments': str(cell(18) or ''),
                }
                if not d['cname_raw'] and not d['confirmation'] and not cell(4):
                    continue
            else:
                if not str(cell(2) or '').strip() and not _to_date(cell(0)) and not cell(5):
                    continue
                ro = 13 if layout == 'fy2627' else 12
                d = {
                    'confirmation': _to_date(cell(0)),
                    'invoice_date': _to_date(cell(1)),
                    'cname_raw': str(cell(2) or '').strip(),
                    'agency_commission': str(cell(3) or '')[:120],
                    'direction_raw': str(cell(4) or ''),
                    'total_fee': _to_decimal(cell(5)),
                    'agency_pct': _to_pct(cell(6)),
                    'agency_inr': _to_decimal(cell(7)),
                    'creator_fee': _to_decimal(cell(8)),
                    'billing_entity': str(cell(9) or '')[:120],
                    'brand': str(cell(10) or '').strip()[:200],
                    'campaign_name': str(cell(11) or ''),
                    'deliverables': str(cell(12) or '')[:255] if layout == 'fy2627' else '',
                    'ro_number': str(cell(ro) or '')[:80],
                    'campaign_over': _to_yn(cell(ro + 1)),
                    'invoice_received': _to_yn(cell(ro + 2)),
                    'payment_cleared': _to_yn(cell(ro + 3)),
                    'e_invoice_no': str(cell(ro + 4) or '')[:80],
                    'payment_received': _to_yn(cell(ro + 5)),
                    'comments': str(cell(ro + 6) or ''),
                }
            parsed.append(d)

        # Inherit brand / campaign across RO sub-row groups (tracker layout):
        # blank cells on /B, /C … rows mean "same campaign as the first row".
        if layout == 'tracker':
            first_of_group: dict[str, dict] = {}
            for d in parsed:
                m = ro_base_re.match(d['ro_number'].strip())
                if not m:
                    continue
                base = _key(m.group(1))
                head = first_of_group.setdefault(base, d)
                if d is head:
                    # A grouped campaign with no recorded name still needs one
                    # row to link the group; derive "<brand> · <Mon YYYY>".
                    if not d['campaign_name'].strip() and d['brand']:
                        anchor = d['invoice_date'] or d['confirmation']
                        suffix = anchor.strftime(' · %b %Y') if anchor else ''
                        d['campaign_name'] = f"{d['brand']}{suffix}"
                    continue
                if not d['brand']:
                    d['brand'] = head['brand']
                if not d['campaign_name'].strip():
                    d['campaign_name'] = head['campaign_name']
                if not d['billing_entity'].strip():
                    d['billing_entity'] = head['billing_entity']
                # One invoice covers the whole group — the workbook's own
                # summary counts sub-row billing under the head's invoice month.
                if not d['e_invoice_no'] and head['e_invoice_no']:
                    d['e_invoice_no'] = head['e_invoice_no']

        count = skipped = 0
        for d in parsed:
            if (_key(d['ro_number']) in existing_ro
                    or (d['e_invoice_no'] and _key(d['e_invoice_no']) in existing_einv)):
                skipped += 1
                continue

            direction_raw = d['direction_raw'].strip().lower()
            direction = (
                'Inbound' if direction_raw == 'inbound' else
                'Outbound' if direction_raw == 'outbound' else
                'MarkUp' if 'mark' in direction_raw else 'Outbound'
            )
            creator = _resolve_creator(d['cname_raw'], default_relationship='NonTCH') if d['cname_raw'] else None

            CommercialDeal.objects.create(
                confirmation_date=d['confirmation'],
                e_invoice_date=d['invoice_date'],
                creator=creator,
                creator_name_raw='' if creator else d['cname_raw'],
                agency_commission_agreed=d['agency_commission'],
                direction=direction,
                total_fee=d['total_fee'],
                agency_fee_pct=d['agency_pct'],
                agency_fee_inr=d['agency_inr'],
                creator_fee=d['creator_fee'],
                billing_entity=d['billing_entity'],
                brand=d['brand'],
                campaign=_resolve_campaign(d['campaign_name'], d['brand']),
                deliverables=d['deliverables'],
                ro_number=d['ro_number'],
                campaign_over=d['campaign_over'],
                invoice_received=d['invoice_received'],
                payment_cleared=d['payment_cleared'],
                e_invoice_number=d['e_invoice_no'],
                payment_received=d['payment_received'],
                comments=d['comments'],
            )
            count += 1

        # Campaign status is derived from its deals' campaign_over flags.
        for campaign in Campaign.objects.all():
            campaign.refresh_status()
        self.stdout.write(
            f'  deals: {count} created, {skipped} skipped as already present  (layout: {layout})'
        )

    def _import_employee_reports(self, ws):
        """Cols: A=SNo, B=WeekEnding, C=Employee, D=Outreach, E=PaidConf,
        F=Revenue, G=Profit, H=Barter, I=LiveCampaigns, J=ActionPoints.
        Week ending is often free text like '21st to 27th Nov' so we tolerate that."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            emp = row[2]
            if not emp:
                continue
            week = _to_date(row[1])
            outreach_raw = row[3]
            try:
                outreach = int(float(outreach_raw)) if outreach_raw not in (None, '') and not isinstance(outreach_raw, str) else 0
            except (ValueError, TypeError):
                outreach = 0
            try:
                live = int(float(row[8])) if row[8] not in (None, '') else 0
            except (ValueError, TypeError):
                live = 0
            EmployeeWeeklyReport.objects.create(
                week_ending=week,
                employee_name=str(emp).strip()[:120],
                new_outreach=outreach,
                paid_confirmations=str(row[4] or '')[:255],
                revenue_locked=_to_decimal(row[5]),
                profit_locked=_to_decimal(row[6]),
                barter_confirmations=str(row[7] or '')[:255],
                live_campaigns=live,
                action_points=str(row[9] or '') if len(row) > 9 else '',
            )
            count += 1
        self.stdout.write(f'  employee reports: {count}')

    def _import_dropoffs(self, ws):
        """Cols: A=SNo, B=Name, C=DropOffDate, D=Reason, E=Learning, F=Duration."""
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[1]
            if not name_raw:
                continue
            name = _norm_name(str(name_raw))
            creator = Creator.objects.filter(name__iexact=name).first()
            d = _to_date(row[2])
            DropOff.objects.create(
                creator=creator,
                creator_name_raw='' if creator else name,
                drop_off_date=d,
                drop_off_date_note='' if d else str(row[2] or '')[:120],
                reason=str(row[3] or ''),
                learning=str(row[4] or ''),
                duration=str(row[5] or '')[:80],
            )
            # If we resolved the creator, also mark them Dropped
            if creator and creator.relationship != 'Dropping':
                creator.relationship = 'Dropping'
                creator.save(update_fields=['relationship'])
            count += 1
        self.stdout.write(f'  drop-offs: {count}')
